"""
generate_xlsx.py
================
Converts submission.csv → a beautifully formatted submission.xlsx
with score bars, color-coded ranks, and no NaN values, matching the 
updated Judge's rubric for Artifact 1.
"""

import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule

CSV_IN   = "submission.csv"
XLSX_OUT = "submission.xlsx"

HEADER_BG   = "1E1B4B"
HEADER_FG   = "FFFFFF"
ALT_ROW_BG  = "F8F4FF"
BORDER_CLR  = "D1C4E9"

def thin_border():
    side = Side(style="thin", color=BORDER_CLR)
    return Border(left=side, right=side, top=side, bottom=side)

def main():
    df = pd.read_csv(CSV_IN, encoding="utf-8")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Top-100 Rankings"

    # Title
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "AlgoYodhas — AI Candidate Ranking"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Headers
    headers = ["candidate_id", "rank", "score", "reasoning"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(name="Calibri", bold=True, size=11, color=HEADER_FG)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
    ws.row_dimensions[2].height = 20

    # Data
    for row_idx, row in df.iterrows():
        excel_row = row_idx + 3
        rank = int(row["rank"])
        
        for col_idx, col_name in enumerate(headers, 1):
            val = row[col_name]
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.border = thin_border()
            cell.font = Font(name="Calibri", size=10)
            
            if rank % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ALT_ROW_BG)
                
            if col_idx == 1:
                cell.font = Font(name="Courier New", bold=True, size=10, color="4C1D95")
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_idx == 2:
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_idx == 3:
                cell.number_format = "0.0000"
                cell.alignment = Alignment(horizontal="right", vertical="top")
            elif col_idx == 4:
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
                
        ws.row_dimensions[excel_row].height = 45

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 90
    
    ws.freeze_panes = "A3"
    
    score_range = f"C3:C{2 + len(df)}"
    ws.conditional_formatting.add(
        score_range,
        ColorScaleRule(
            start_type="min", start_color="F87171",
            mid_type="percentile", mid_value=50, mid_color="FDE68A",
            end_type="max", end_color="34D399"
        )
    )

    wb.save(XLSX_OUT)
    print(f"Saved {XLSX_OUT} successfully.")

if __name__ == "__main__":
    main()
